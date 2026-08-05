#!/usr/bin/env python3
"""
Sincronização ERP (planilha) -> Supabase - Sistema Wallace Lira
======================================================================
NOVO 04/08/2026 (parte 84, pedido explícito do usuário: "não pode em nenhuma
hipótese estar sem sincronismo entre o site e o ERP, o ERP deve ser a verdade
absoluta... não deve sair alterando 10 números iguais no site, só deve mudar
em 1 lugar, o site que busque a fonte").

O QUE ISSO RESOLVE:
  Antes desta sessão, HISTORICO_ERP_TODOS_CICLOS (usado pela reconciliação
  Pluggy/Inbox e por outras partes do site) era extraído manualmente da
  planilha por sessões de chat e colado direto no app.js/Supabase - sem
  automação nenhuma. Isso já causou pelo menos 1 divergência real (correção
  da parte 71 "sumindo" entre sessões porque não existia fonte única).

  Este script fecha esse buraco: lê a aba SWP_INPUT_TX do ERP (.xlsx) de
  verdade e escreve o resultado DIRETO no Supabase (wallace_dados.dados ->
  HISTORICO_ERP_TODOS_CICLOS), sem intervenção manual. Rodar isso sempre que
  a planilha for atualizada (ou agendar, mesmo padrão do sincronizar_pluggy.py)
  garante que o site nunca mais fica dessincronizado do ERP.

IMPORTANTE - o que este script NÃO faz:
  Não mexe em nenhum outro dado do site (saldos de caixa, agregados como
  cartaoMBTotal, etc.) - só no HISTORICO_ERP_TODOS_CICLOS, que é
  especificamente "todos os lançamentos de todos os ciclos, pra
  reconciliação". Os 7 livros "vivos" do ciclo atual (LRW_TRANSACOES etc)
  continuam sendo mantidos à parte, como já eram.

USO:
  python3 sincronizar_erp_supabase.py caminho/para/ERP_WALLACE_LIRA_V11.xlsx

Variáveis de ambiente necessárias:
  SUPABASE_URL  - https://bakdgacmwlopvrrppwdm.supabase.co
  SUPABASE_KEY  - chave do Supabase (mesma dos outros scripts)
"""
import json
import os
import sys
from datetime import datetime, timedelta
from urllib.request import Request, urlopen
from urllib.error import HTTPError, URLError

try:
    import openpyxl
except ImportError:
    print("ERRO: openpyxl não instalado. Rode: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def excel_date_to_str(serial):
    """Converte data serial do Excel (base 1899-12-30, padrão do Windows Excel)
    pra string DD/MM/AAAA. Retorna None se o valor não for um serial válido."""
    if serial is None:
        return None
    try:
        serial = float(serial)
    except (TypeError, ValueError):
        return None
    try:
        dt = datetime(1899, 12, 30) + timedelta(days=serial)
        return dt.strftime("%d/%m/%Y")
    except (OverflowError, ValueError):
        return None


def extrair_historico(caminho_xlsx: str) -> list[dict]:
    """Lê a aba SWP_INPUT_TX inteira, devolve todos os lançamentos com VALOR
    preenchido, qualquer livro, qualquer ciclo - mesmo criterio já usado
    manualmente nas sessões anteriores (04/08/2026), agora automatizado."""
    wb = openpyxl.load_workbook(caminho_xlsx, data_only=True)
    if "SWP_INPUT_TX" not in wb.sheetnames:
        raise RuntimeError(f"Aba SWP_INPUT_TX não encontrada em {caminho_xlsx} - planilha errada ou versão antiga (V10 não tem essa aba, precisa da V11+).")
    ws = wb["SWP_INPUT_TX"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise RuntimeError("Aba SWP_INPUT_TX está vazia.")
    header = rows[0]
    idx = {h: i for i, h in enumerate(header) if h is not None}

    obrigatorias = ["TX", "DATA", "LIVRO", "VALOR"]
    faltando = [c for c in obrigatorias if c not in idx]
    if faltando:
        raise RuntimeError(f"Colunas obrigatórias faltando em SWP_INPUT_TX: {faltando}")

    itens = []
    for r in rows[1:]:
        tx = r[idx["TX"]]
        if not tx:
            continue
        valor = r[idx["VALOR"]]
        if valor is None:
            continue
        try:
            v = round(abs(float(valor)), 2)
        except (TypeError, ValueError):
            continue
        data_str = excel_date_to_str(r[idx["DATA"]])
        livro = r[idx["LIVRO"]] or "?"
        nome_col = idx.get("ESTABELECIMENTO")
        desc_col = idx.get("DESCRICAO")
        nome = None
        if nome_col is not None:
            nome = r[nome_col]
        if not nome and desc_col is not None:
            nome = r[desc_col]
        itens.append({
            "tx": str(tx),
            "data": data_str or "",
            "livro": str(livro),
            "nome": str(nome)[:60] if nome else "",
            "valor": v,
        })
    return itens


def atualizar_supabase(supabase_url: str, supabase_key: str, itens: list[dict]) -> None:
    headers = {
        "apikey": supabase_key,
        "Authorization": f"Bearer {supabase_key}",
        "Content-Type": "application/json",
    }
    # Escreve direto na chave de topo HISTORICO_ERP_TODOS_CICLOS via PATCH no /rest/v1 -
    # equivalente ao "dados = dados || jsonb_build_object(...)" já usado manualmente nas
    # sessões anteriores, agora automatizado. PATCH com Prefer: merge-duplicates não se aplica
    # a JSONB assim, então lemos o registro, mesclamos em memória, e regravamos - mais simples
    # e seguro que tentar fazer o merge no lado do Postgres via REST puro.
    url_get = f"{supabase_url}/rest/v1/wallace_dados?select=dados&id=eq.1"
    req_get = Request(url_get, headers=headers, method="GET")
    with urlopen(req_get, timeout=20) as resp:
        linhas = json.loads(resp.read().decode("utf-8"))
    dados_atual = linhas[0]["dados"] if linhas else {}
    dados_atual["HISTORICO_ERP_TODOS_CICLOS"] = itens

    url_patch = f"{supabase_url}/rest/v1/wallace_dados?id=eq.1"
    body = json.dumps({"dados": dados_atual}).encode("utf-8")
    req_patch = Request(url_patch, data=body, headers=headers, method="PATCH")
    with urlopen(req_patch, timeout=20) as resp:
        resp.read()
    print(f"Supabase atualizado: {len(itens)} lançamentos gravados em HISTORICO_ERP_TODOS_CICLOS.")


def main() -> int:
    if len(sys.argv) < 2:
        print("USO: python3 sincronizar_erp_supabase.py caminho/para/ERP_WALLACE_LIRA_V11.xlsx", file=sys.stderr)
        return 1
    caminho_xlsx = sys.argv[1]
    if not os.path.isfile(caminho_xlsx):
        print(f"ERRO: arquivo não encontrado: {caminho_xlsx}", file=sys.stderr)
        return 1

    supabase_url = os.environ.get("SUPABASE_URL")
    supabase_key = os.environ.get("SUPABASE_KEY")
    if not supabase_url or not supabase_key:
        print("ERRO: SUPABASE_URL e/ou SUPABASE_KEY não definidos.", file=sys.stderr)
        return 1

    try:
        print(f"Lendo {caminho_xlsx}...")
        itens = extrair_historico(caminho_xlsx)
        print(f"{len(itens)} lançamentos encontrados (todos os livros, todos os ciclos).")

        por_livro: dict[str, int] = {}
        for it in itens:
            por_livro[it["livro"]] = por_livro.get(it["livro"], 0) + 1
        for livro, qtd in sorted(por_livro.items(), key=lambda x: -x[1]):
            print(f"  {livro}: {qtd} lançamento(s)")

        print("\nAtualizando Supabase...")
        atualizar_supabase(supabase_url, supabase_key, itens)
        print("\nFeito. O site vai ler esses dados na próxima carga de página, sem precisar de deploy.")
        return 0
    except (RuntimeError, HTTPError, URLError) as e:
        print(f"ERRO: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
