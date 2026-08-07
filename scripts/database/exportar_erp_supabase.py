#!/usr/bin/env python3
"""
Exportação Supabase -> Excel - Sistema Wallace Lira (Fase 4 da Arquitetura V2)
======================================================================
NOVO 05/08/2026 (parte 104). Inverte a direção do antigo
`sincronizar_erp_supabase.py`: agora que a tabela relacional `transacoes`
existe de verdade no Supabase (Fase 1-2 da Arquitetura V2, 281 lançamentos
migrados e auditados), ELA é a fonte de verdade - não a planilha.

O QUE ISSO RESOLVE:
  Antes, a via de entrada era planilha -> Supabase (sincronizar_erp_supabase.py).
  Agora que existe uma tabela relacional com lançamento direto (via Claude/MCP
  ou futura tela de lançamento manual), a planilha vira só uma exportação de
  leitura - útil pra auditoria humana (mais fácil abrir uma aba do Excel que
  rodar SQL), mas nunca mais fonte de escrita.

IMPORTANTE - o que este script NÃO faz:
  Não escreve nada de volta no Supabase. Não lê nenhuma planilha existente.
  Rodar isso não tem efeito nenhum no site - é só uma cópia pra você conferir.

USO:
  python3 exportar_erp_supabase.py caminho/de/saida/ERP_EXPORTADO.xlsx

Variáveis de ambiente necessárias:
  SUPABASE_URL  - https://bakdgacmwlopvrrppwdm.supabase.co
  SUPABASE_KEY  - chave do Supabase (mesma dos outros scripts)
"""
import json
import os
import sys
from urllib.request import Request, urlopen
from urllib.error import HTTPError, URLError

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERRO: openpyxl não instalado. Rode: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def buscar_rest(supabase_url: str, supabase_key: str, tabela: str, select: str, ordem: str = "") -> list[dict]:
    """GET genérico no PostgREST. `ordem`, se dado, deve vir pronto (ex: 'data.asc')."""
    headers = {"apikey": supabase_key, "Authorization": f"Bearer {supabase_key}"}
    url = f"{supabase_url}/rest/v1/{tabela}?select={select}"
    if ordem:
        url += f"&order={ordem}"
    todos: list[dict] = []
    offset = 0
    tamanho_pagina = 1000
    while True:
        req_headers = dict(headers)
        req_headers["Range-Unit"] = "items"
        req_headers["Range"] = f"{offset}-{offset + tamanho_pagina - 1}"
        req = Request(url, headers=req_headers, method="GET")
        with urlopen(req, timeout=30) as resp:
            pagina = json.loads(resp.read().decode("utf-8"))
        todos.extend(pagina)
        if len(pagina) < tamanho_pagina:
            break
        offset += tamanho_pagina
    return todos


def montar_workbook(transacoes: list[dict], caixas: dict[str, str], categorias: dict[str, str]) -> openpyxl.Workbook:
    """Recria o formato SWP_INPUT_TX (TX, DATA, LIVRO, VALOR, ESTABELECIMENTO, TIPO, STATUS,
    CATEGORIA) pra quem já está acostumado a auditar nessa aba - só que agora é sempre um
    espelho do Supabase, nunca editado à mão de volta."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SWP_INPUT_TX (exportado)"

    cabecalho = ["TX", "DATA", "LIVRO (caixa)", "TIPO", "VALOR", "ESTABELECIMENTO", "CATEGORIA", "STATUS", "ORIGEM"]
    for col, titulo in enumerate(cabecalho, 1):
        c = ws.cell(row=1, column=col, value=titulo)
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F1F1F")

    for r, t in enumerate(transacoes, 2):
        ws.cell(row=r, column=1, value=t.get("tx_legado") or t.get("id"))
        ws.cell(row=r, column=2, value=t.get("data"))
        ws.cell(row=r, column=3, value=caixas.get(t.get("caixa_id"), ""))
        ws.cell(row=r, column=4, value=t.get("tipo"))
        v = ws.cell(row=r, column=5, value=float(t["valor"]) if t.get("valor") is not None else None)
        v.number_format = "#,##0.00"
        ws.cell(row=r, column=6, value=t.get("descricao"))
        ws.cell(row=r, column=7, value=categorias.get(t.get("categoria_id"), ""))
        ws.cell(row=r, column=8, value=t.get("status"))
        ws.cell(row=r, column=9, value=t.get("origem"))
        for col in range(1, 10):
            ws.cell(row=r, column=col).font = Font(name="Arial", size=10)

    larguras = [24, 14, 22, 10, 12, 46, 20, 20, 16]
    for i, w in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{len(transacoes) + 1}"

    aviso = wb.create_sheet("LEIA-ME")
    aviso["A1"] = "Este arquivo é uma EXPORTAÇÃO somente-leitura do Supabase (fonte de verdade)."
    aviso["A2"] = "Editar aqui NÃO tem efeito nenhum no site - a via de entrada agora é o Supabase (via Claude/MCP)."
    aviso["A3"] = f"Gerado em: {__import__('datetime').datetime.now().strftime('%d/%m/%Y %H:%M')}"
    aviso["A4"] = f"Total de transações exportadas: {len(transacoes)}"
    for row in range(1, 5):
        aviso.cell(row=row, column=1).font = Font(name="Arial", size=11)
    aviso.column_dimensions["A"].width = 90

    return wb


def main() -> int:
    if len(sys.argv) < 2:
        print("USO: python3 exportar_erp_supabase.py caminho/de/saida/ERP_EXPORTADO.xlsx", file=sys.stderr)
        return 1
    caminho_saida = sys.argv[1]

    supabase_url = os.environ.get("SUPABASE_URL")
    supabase_key = os.environ.get("SUPABASE_KEY")
    if not supabase_url or not supabase_key:
        print("ERRO: SUPABASE_URL e/ou SUPABASE_KEY não definidos.", file=sys.stderr)
        return 1

    try:
        print("Lendo caixas e categorias (pra traduzir os IDs)...")
        caixas_raw = buscar_rest(supabase_url, supabase_key, "caixas", "id,nome")
        categorias_raw = buscar_rest(supabase_url, supabase_key, "categorias", "id,nome")
        caixas = {c["id"]: c["nome"] for c in caixas_raw}
        categorias = {c["id"]: c["nome"] for c in categorias_raw}

        print("Lendo transacoes...")
        transacoes = buscar_rest(
            supabase_url, supabase_key, "transacoes",
            "id,tx_legado,data,tipo,valor,descricao,caixa_id,categoria_id,status,origem",
            ordem="data.asc",
        )
        print(f"{len(transacoes)} transações encontradas.")

        wb = montar_workbook(transacoes, caixas, categorias)
        wb.save(caminho_saida)
        print(f"\nExportado: {caminho_saida}")
        print("Lembrete: este arquivo é só-leitura. A fonte de verdade continua sendo o Supabase.")
        return 0
    except (HTTPError, URLError) as e:
        print(f"ERRO de rede/Supabase: {e}", file=sys.stderr)
        return 1
    except Exception as e:
        print(f"ERRO: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
