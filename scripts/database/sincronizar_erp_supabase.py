#!/usr/bin/env python3
"""
Sincronização ERP (planilha) -> Supabase - Sistema Wallace Lira
======================================================================
NOVO 04/08/2026 (parte 84, pedido explícito do usuário: "não pode em nenhuma
hipótese estar sem sincronismo entre o site e o ERP, o ERP deve ser a verdade
absoluta... não deve sair alterando 10 números iguais no site, só deve mudar
em 1 lugar, o site que busque a fonte").

MIGRADO 10/08/2026 (pedido explícito do usuário: "esqueça V1, ele está morto,
tudo deve refletir V2" + "faça o trabalho completo" — fila original desta
sessão "demais leitores/escritores ativos da V1"): este script escrevia em
wallace_dados.dados.HISTORICO_ERP_TODOS_CICLOS (V1). A leitura já preferia V2
há dias (view vw_historico_erp_completo, ver app.js) — agora a ESCRITA
também vai direto pra tabela relacional `transacoes` (tx_legado = coluna TX
da planilha), sem tocar mais em wallace_dados. Cada lançamento é gravado com
afeta_saldo_real=false (não mexe em nenhum saldo de caixa atual - histórico é
só pra reconciliação/busca, mesma regra de sempre) e status='confirmado'.
UPSERT idempotente via ON CONFLICT (tx_legado, caixa_id) DO NOTHING - rodar
o script 2x com a mesma planilha não duplica nada.

O QUE ISSO RESOLVE:
  HISTORICO_ERP_TODOS_CICLOS (usado pela reconciliação Pluggy/Inbox e pela
  Busca Global) era extraído manualmente da planilha por sessões de chat e
  colado direto no Supabase - sem automação nenhuma. Isso já causou pelo
  menos 1 divergência real (correção "sumindo" entre sessões porque não
  existia fonte única). Este script fecha esse buraco.

IMPORTANTE - o que este script NÃO faz:
  Não mexe em nenhum outro dado do site (saldos de caixa, agregados como
  cartaoMBTotal, etc.) - só grava histórico com afeta_saldo_real=false,
  que nenhuma view de saldo soma. Os 7 livros "vivos" do ciclo atual
  (LRW_TRANSACOES etc) continuam sendo mantidos à parte, como já eram.

MAPEAMENTO LIVRO -> CAIXA (ver MAPA_LIVRO_CAIXA_ID abaixo):
  `transacoes.caixa_id` é obrigatório (NOT NULL) mas a planilha só tem o
  código do livro em texto (ex: "LRW"). Os livros de cartão (LRW/LRV e
  variantes, LRP, LRS, LRR - todos ligados a Visa Infinite/Mastercard Black)
  mapeiam pra "Caixa Mastercard/Infinite" (mesma caixa que paga a fatura) -
  não atribui cartao_id específico (isso reabriria a classificação de
  Cartões, fora de escopo desta migração - fica NULL, sem problema, coluna
  é opcional). Livros sem caixa clara (LRC "limbo", LRCON, P2P) caem no
  fallback "Caixa Variável" (caixa operacional mais genérica).

USO:
  python3 sincronizar_erp_supabase.py caminho/para/ERP_WALLACE_LIRA_V11.xlsx

Variáveis de ambiente necessárias:
  SUPABASE_URL  - https://bakdgacmwlopvrrppwdm.supabase.co
  SUPABASE_KEY  - chave do Supabase com permissão de INSERT em `transacoes`
"""
import json
import os
import sys
from datetime import datetime, timedelta
from urllib.request import Request, urlopen
from urllib.error import HTTPError, URLError

try:
    import openpyxl
except ImportError:
    print("ERRO: openpyxl não instalado. Rode: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def excel_date_to_str(serial):
    """Converte data serial do Excel (base 1899-12-30, padrão do Windows Excel)
    pra string DD/MM/AAAA. Retorna None se o valor não for um serial válido."""
    if serial is None:
        return None
    try:
        serial = float(serial)
    except (TypeError, ValueError):
        return None
    try:
        dt = datetime(1899, 12, 30) + timedelta(days=serial)
        return dt.strftime("%d/%m/%Y")
    except (OverflowError, ValueError):
        return None


# Mapeamento LIVRO (coluna da planilha) -> caixas.id (V2). Ver nota no docstring do topo do
# arquivo - livros de cartão (Visa/MB) mapeiam pra "Caixa Mastercard/Infinite" (sem cartao_id
# específico, fora de escopo); livros sem caixa clara caem em "Caixa Variável".
_CAIXA_MASTERCARD_INFINITE = "748b8612-b854-44e3-8834-542ec7f1ff7c"
_CAIXA_VARIAVEL = "8522e256-2039-4c11-bd28-69738bfcf5b8"
MAPA_LIVRO_CAIXA_ID = {
    "LRW": _CAIXA_MASTERCARD_INFINITE, "LRW-I": _CAIXA_MASTERCARD_INFINITE, "LRW-MB": _CAIXA_MASTERCARD_INFINITE,
    "LRV": _CAIXA_MASTERCARD_INFINITE, "LRV-I": _CAIXA_MASTERCARD_INFINITE, "LRV-MB": _CAIXA_MASTERCARD_INFINITE,
    "LRV_HISTORICO": _CAIXA_MASTERCARD_INFINITE,
    "LRP": _CAIXA_MASTERCARD_INFINITE,   # Parcelamentos Visa
    "LRS": _CAIXA_MASTERCARD_INFINITE,   # Assinaturas (Mastercard Black)
    "LRR": _CAIXA_MASTERCARD_INFINITE,   # Recorrências (Mastercard Black)
    "LRCV": _CAIXA_VARIAVEL,
    "LRPG": "fb779cdc-ab92-492d-a172-8d147d1380ea",   # PIX Geral Vanessa
    "LRMP": "7ddff812-d54a-4df6-b6bf-ae3351c9fcfe",   # Mercado Pago
    "LRB":  "7751575a-6339-4bf2-bda4-60817778551c",   # Caixa Boletos
    "LRC":  _CAIXA_VARIAVEL,             # "limbo" - sem caixa própria, fallback
    "LRPV": "6c6546fa-5b83-4db6-aa33-ac1bf35370d9",   # PIX Vanessa
    "LRCON": _CAIXA_VARIAVEL,            # Consórcio - sem caixa própria, fallback
    "P2P": _CAIXA_VARIAVEL,              # sem caixa própria, fallback
    "LRSF": "d15e8cbe-4443-4ee4-9631-06d8d49058fe",   # Caixa Saúde Família
    "LRCL": "ff0cd9af-c5a9-4a9b-8cdd-c379e167275e",   # Caixa Lance
}


def extrair_historico(caminho_xlsx: str) -> list[dict]:
    """Lê a aba SWP_INPUT_TX inteira, devolve todos os lançamentos com VALOR
    preenchido, qualquer livro, qualquer ciclo - mesmo criterio já usado
    manualmente nas sessões anteriores (04/08/2026), agora automatizado.
    ATUALIZADO 10/08/2026: também extrai TIPO (Entrada/Saida) e mapeia LIVRO
    pra caixa_id (MAPA_LIVRO_CAIXA_ID) - linhas com livro sem mapeamento
    conhecido são puladas com aviso, nunca gravadas com caixa_id chutado."""
    wb = openpyxl.load_workbook(caminho_xlsx, data_only=True)
    if "SWP_INPUT_TX" not in wb.sheetnames:
        raise RuntimeError(f"Aba SWP_INPUT_TX não encontrada em {caminho_xlsx} - planilha errada ou versão antiga (V10 não tem essa aba, precisa da V11+).")
    ws = wb["SWP_INPUT_TX"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise RuntimeError("Aba SWP_INPUT_TX está vazia.")
    header = rows[0]
    idx = {h: i for i, h in enumerate(header) if h is not None}

    obrigatorias = ["TX", "DATA", "LIVRO", "TIPO", "VALOR"]
    faltando = [c for c in obrigatorias if c not in idx]
    if faltando:
        raise RuntimeError(f"Colunas obrigatórias faltando em SWP_INPUT_TX: {faltando}")

    itens = []
    sem_mapa = set()
    for r in rows[1:]:
        tx = r[idx["TX"]]
        if not tx:
            continue
        valor = r[idx["VALOR"]]
        if valor is None:
            continue
        try:
            v = round(abs(float(valor)), 2)
        except (TypeError, ValueError):
            continue
        if v == 0:
            continue
        data_raw = r[idx["DATA"]]
        if isinstance(data_raw, datetime):
            data_iso = data_raw.strftime("%Y-%m-%d")
        else:
            data_str = excel_date_to_str(data_raw)
            data_iso = None
            if data_str:
                dia, mes, ano = data_str.split("/")
                data_iso = f"{ano}-{mes}-{dia}"
        livro = str(r[idx["LIVRO"]] or "").strip()
        caixa_id = MAPA_LIVRO_CAIXA_ID.get(livro)
        if not caixa_id:
            sem_mapa.add(livro)
            continue
        tipo_raw = str(r[idx["TIPO"]] or "").strip().lower()
        tipo = "entrada" if tipo_raw == "entrada" else "saida"
        nome_col = idx.get("ESTABELECIMENTO")
        desc_col = idx.get("DESCRICAO")
        nome = None
        if nome_col is not None:
            nome = r[nome_col]
        if not nome and desc_col is not None:
            nome = r[desc_col]
        itens.append({
            "tx": str(tx),
            "data": data_iso,
            "livro": livro,
            "caixa_id": caixa_id,
            "tipo": tipo,
            "nome": str(nome)[:60] if nome else "",
            "valor": v,
        })
    if sem_mapa:
        print(f"AVISO: {len(sem_mapa)} código(s) de LIVRO sem mapeamento conhecido, linhas puladas: {sorted(sem_mapa)}. Adicionar em MAPA_LIVRO_CAIXA_ID se forem reais.", file=sys.stderr)
    return itens


def atualizar_supabase(supabase_url: str, supabase_key: str, itens: list[dict]) -> None:
    """MIGRADO 10/08/2026: grava direto na tabela relacional `transacoes` (V2), não mais em
    wallace_dados (V1) - ver docstring do topo do arquivo. UPSERT via on_conflict pro par
    (tx_legado, caixa_id) - mesmo padrão já usado em outros scripts deste projeto (ex:
    atualizar_geracao_saj.py, on_conflict=data). afeta_saldo_real=false SEMPRE - histórico é só
    pra reconciliação/busca, nunca deve mudar nenhum saldo de caixa atual."""
    headers = {
        "apikey": supabase_key,
        "Authorization": f"Bearer {supabase_key}",
        "Content-Type": "application/json",
        "Prefer": "resolution=ignore-duplicates,return=minimal",
    }
    url = f"{supabase_url}/rest/v1/transacoes?on_conflict=tx_legado,caixa_id"
    body_itens = [{
        "data": it["data"],
        "descricao": it["nome"] or it["livro"],
        "valor": it["valor"],
        "tipo": it["tipo"],
        "caixa_id": it["caixa_id"],
        "origem": "reconciliacao",
        "status": "confirmado",
        "afeta_saldo_real": False,
        "tx_legado": it["tx"],
    } for it in itens]
    # Em lotes de 100 - corpo grande demais numa unica requisicao arrisca timeout/limite do PostgREST.
    TAMANHO_LOTE = 100
    for i in range(0, len(body_itens), TAMANHO_LOTE):
        lote = body_itens[i:i + TAMANHO_LOTE]
        body = json.dumps(lote).encode("utf-8")
        req = Request(url, data=body, headers=headers, method="POST")
        with urlopen(req, timeout=30) as resp:
            resp.read()
    print(f"Supabase (V2, tabela transacoes) atualizado: {len(itens)} lançamentos enviados (novos gravados, já existentes ignorados via ON CONFLICT).")


def main() -> int:
    if len(sys.argv) < 2:
        print("USO: python3 sincronizar_erp_supabase.py caminho/para/ERP_WALLACE_LIRA_V11.xlsx", file=sys.stderr)
        return 1
    caminho_xlsx = sys.argv[1]
    if not os.path.isfile(caminho_xlsx):
        print(f"ERRO: arquivo não encontrado: {caminho_xlsx}", file=sys.stderr)
        return 1

    supabase_url = os.environ.get("SUPABASE_URL")
    supabase_key = os.environ.get("SUPABASE_KEY")
    if not supabase_url or not supabase_key:
        print("ERRO: SUPABASE_URL e/ou SUPABASE_KEY não definidos.", file=sys.stderr)
        return 1

    try:
        print(f"Lendo {caminho_xlsx}...")
        itens = extrair_historico(caminho_xlsx)
        print(f"{len(itens)} lançamentos encontrados (todos os livros, todos os ciclos).")

        por_livro: dict[str, int] = {}
        for it in itens:
            por_livro[it["livro"]] = por_livro.get(it["livro"], 0) + 1
        for livro, qtd in sorted(por_livro.items(), key=lambda x: -x[1]):
            print(f"  {livro}: {qtd} lançamento(s)")

        print("\nAtualizando Supabase...")
        atualizar_supabase(supabase_url, supabase_key, itens)
        print("\nFeito. O site vai ler esses dados na próxima carga de página, sem precisar de deploy.")
        return 0
    except (RuntimeError, HTTPError, URLError) as e:
        print(f"ERRO: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
